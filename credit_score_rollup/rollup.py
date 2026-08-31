#!/usr/bin/env python3
"""Stack credit-band tabs in a workbook into combined "_ALL" tabs.

Usage:
    python rollup.py BOOK.xlsx                 # writes BOOK.rollup.xlsx
    python rollup.py BOOK.xlsx -o OUT.xlsx
    python rollup.py BOOK.xlsx --in-place
    python rollup.py BOOK.xlsx --no-band-column
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from rollup_config import ROLLUPS

BAND_COLUMN = "Credit Band"


def normalize(name):
    """Fold a sheet name so cosmetic differences don't break matching."""
    return re.sub(r"[\s_\-]+", "", str(name)).lower()


def band_of(sheet_name):
    """The credit band a source sheet covers, for the added label column."""
    folded = normalize(sheet_name)
    if folded.endswith("640+"):
        return "640+"
    if folded.endswith("601639"):
        return "601-639"
    if folded.endswith("600"):
        return "<=600"
    return sheet_name


def resolve(workbook, wanted):
    """Map wanted sheet names onto the workbook's actual sheet names."""
    actual = {normalize(name): name for name in workbook.sheetnames}
    found, missing = [], []
    for name in wanted:
        match = actual.get(normalize(name))
        (found if match else missing).append(match or name)
    return found, missing


def rows_of(sheet):
    """Non-empty rows of a sheet as lists of values."""
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            yield list(row)


def build_sheet(workbook, target, sources, band_column):
    """Replace `target` with the stacked contents of `sources`."""
    if target in workbook.sheetnames:
        del workbook[target]
    out = workbook.create_sheet(target)

    header = None
    written = 0
    for name in sources:
        rows = list(rows_of(workbook[name]))
        if not rows:
            print(f"    {name}: empty, skipped")
            continue
        band = band_of(name)
        if header is None:
            header = rows[0]
            out.append(header + ([BAND_COLUMN] if band_column else []))
        elif normalize("".join(str(c) for c in rows[0])) != normalize(
            "".join(str(c) for c in header)
        ):
            print(f"    {name}: WARNING header differs from first sheet")
        for row in rows[1:]:
            out.append(row + ([band] if band_column else []))
            written += 1
        print(f"    {name}: {len(rows) - 1} rows")

    if header:
        for cell in out[1]:
            cell.font = Font(bold=True)
        out.freeze_panes = "A2"
        widths = [0] * len(out[1])
        for row in out.iter_rows(values_only=True):
            for i, value in enumerate(row):
                widths[i] = max(widths[i], len(str(value)) if value is not None else 0)
        for i, width in enumerate(widths, start=1):
            out.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 40)
    return written


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--in-place", action="store_true")
    parser.add_argument(
        "--no-band-column",
        action="store_true",
        help=f"don't append a '{BAND_COLUMN}' column identifying each row's source tab",
    )
    args = parser.parse_args(argv)

    if args.output and args.in_place:
        parser.error("--output and --in-place are mutually exclusive")
    out_path = args.output or (
        args.workbook if args.in_place
        else args.workbook.with_suffix(f".rollup{args.workbook.suffix}")
    )

    # data_only so cached formula results are stacked, not the formulas, which
    # would point at the wrong cells once rows are restacked.
    workbook = load_workbook(args.workbook, data_only=True)

    total, built = 0, 0
    for target, sources in ROLLUPS.items():
        found, missing = resolve(workbook, sources)
        print(f"{target}:")
        for name in missing:
            print(f"    {name}: NOT FOUND")
        if not found:
            print("    no source tabs present, skipped")
            continue
        total += build_sheet(workbook, target, found, not args.no_band_column)
        built += 1

    if not built:
        print("\nNo roll-ups built - no matching source tabs in this workbook.")
        print("Sheets present: " + ", ".join(workbook.sheetnames))
        return 1

    workbook.save(out_path)
    print(f"\n{built} roll-up tab(s), {total} data rows -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
